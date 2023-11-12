import yfinance as yf
from openpyxl import load_workbook
from datetime import datetime
from tabulate import tabulate
from colorama import Fore
import pandas as pd
import traceback as tb
from utilities.contants import PRINT_DETAILED_EXCEPTION, PRINT_INTERMEDIATE, DEEP_DEBUG


def write_current_sheet(filepath, df):
    # +-----------------+------------------------------------------------------------------+
    # | Definition      | Write all the attributes values for all the tickers for the      |
    # |                 | given date to the 'Stock_Current' sheet                          |
    # +-----------------+------------------------------------------------------------------+
    # | Parameters      | filepath - Path (relative) of the input file                     |
    # |                 | df - Pandas dataframe containing data for all the attributes for |
    # |                 | all the tickers for the given date                               |
    # +-----------------+------------------------------------------------------------------+
    # | Returns         | None                                                             |
    # +-----------------+------------------------------------------------------------------+

    with pd.ExcelWriter(filepath, mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name='Stock_Current', index=False, header=False, startrow=1)
    print(Fore.GREEN, '\nStock data for all tickers saved successfully in Stock_Current sheet')


def write_history_sheet(filepath, ws, df):
    # +-----------------+------------------------------------------------------------------+
    # | Definition      | Append all the attributes values for all the tickers for the     |
    # |                 | given date to the existing data of 'Stock_History' sheet         |                  |
    # +-----------------+------------------------------------------------------------------+
    # | Parameters      | filepath - Path (relative) of the input file                     |
    # |                 | ws - Reference to the worksheet 'Stock_History'                  |
    # |                 | df - Pandas dataframe containing data for all the attributes for |
    # |                 | all the tickers for the given date                               |
    # +-----------------+------------------------------------------------------------------+
    # | Returns         | None                                                             |
    # +-----------------+------------------------------------------------------------------+

    # Get the maximum rowcount of existing data
    startrow = ws.max_row

    # Write data from the next row onwards
    with pd.ExcelWriter(filepath, mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name='Stock_History', index=False, header=False, startrow=startrow)
    print(Fore.GREEN, '\nStock data for all tickers saved successfully in Stock_History sheet')


def fetch_attribs_data(ws, dt):
    # +-----------------+------------------------------------------------------------------+
    # | Definition      | Fetch all the attributes values for all the tickers mentioned    |
    # |                 | in the 'Stock_Current' sheet, for the given date                 |
    # +-----------------+------------------------------------------------------------------+
    # | Parameters      | ws - Reference to 'Stock_Current'                                |
    # |                 | dt - Given date to fetch the stock attributes                    |
    # +-----------------+------------------------------------------------------------------+
    # | Returns         | A dictionary containing the list of attributes and all the       |
    # |                 | attribute values for all the tickers for the given date          |
    # +-----------------+------------------------------------------------------------------+

    # Get total number of tickers and attributes
    num_tckrs = ws.max_row
    num_attribs = ws.max_column
    if PRINT_INTERMEDIATE:
        print('Number of Tickers : ' + str(num_tckrs) if not DEEP_DEBUG else 'Number of Tickers (num_tckrs): ' + str(num_tckrs))
        print('Number of Attributes : ' + str(num_attribs) if not DEEP_DEBUG else 'Number of Attributes (num_attribs): ' + str(num_attribs))

    l_attribs = []

    # Create list of attributes
    for i in range(2, num_attribs + 1):
        l_attribs.append(ws.cell(1, i).value.strip())

    print('Attribute list : ' + str(l_attribs) if not DEEP_DEBUG else 'Attribute list (l_attribs): ' + str(l_attribs))

    l_master_tckr_data = []
    d_master_data = {'data': [],
                     'columns': []}

    # Fetch attribute data for each ticker
    for rec_ctr in range(2, num_tckrs + 1):
        tckr = ws.cell(rec_ctr, 1).value
        print('Processing Ticker ' + tckr if not DEEP_DEBUG else 'Processing Ticker (tckr): ' + tckr)

        # Create ticker object
        feed = yf.Ticker(tckr)
        # Get history feed for ticker
        df_hist = feed.history(period="max").reset_index()
        # Cast Date field to datetype
        df_hist['Date'] = df_hist['Date'].dt.date
        # Filter history feed for given date
        df_hist = df_hist[df_hist.Date == dt]
        if PRINT_INTERMEDIATE:
            print(tabulate(df_hist, headers='keys', tablefmt='psql'))

        # Access data only if dataframe not empty
        if not df_hist.empty:
            l_current_tckr_data = [tckr]
            # Round off all data to 2 decimal places, except for Date
            for attrib in l_attribs:
                l_current_tckr_data.append(round(df_hist[attrib].iloc[0], 2) if attrib != 'Date' else df_hist[attrib].iloc[0])

            l_master_tckr_data.append(l_current_tckr_data)


    if PRINT_INTERMEDIATE:
        print('Ticker data : \n' + tabulate(list(l_master_tckr_data), headers='keys', tablefmt='psql') if not DEEP_DEBUG else 'Ticker data (l_master_tckr_data): \n' + tabulate(list(l_master_tckr_data), headers='keys', tablefmt='psql'))

    d_master_data['data'] = l_master_tckr_data
    d_master_data['columns'] = ['Ticker'] + l_attribs

    return d_master_data



if __name__ == '__main__':

    input_path = "input/"
    input_filename = "stocks.xlsx"
    input_file_abs_path = input_path + input_filename

    try:
        # Load input sheets
        wb = load_workbook(filename = input_file_abs_path)
        ws_current = wb['Stock_Current']
        ws_history = wb['Stock_History']

        # Get date to be used for fetching stock data as user input
        dt = input('Enter Date to extract data: ')
        dt = datetime.strptime(dt, "%d/%m/%y").date()

        # Populate Stock_Current sheet with data as per given date as parameter
        master_data = fetch_attribs_data(ws_current, dt)
        if PRINT_INTERMEDIATE:
            print('Master Data: \n' + tabulate(master_data, headers='keys', tablefmt='psql') if not DEEP_DEBUG else 'Master Data (master_data): \n' + tabulate(master_data, headers='keys', tablefmt='psql'))

        # Trigger data writing only if stock data for all the tickers are not empty
        if master_data['data']:
            df_master_data = pd.DataFrame(master_data['data'], columns=master_data['columns'])

            # Save data to 'Stock_Current' sheet
            write_current_sheet(input_file_abs_path, df_master_data)

            # Append the stock data to the existing data of Stock_History sheet
            write_history_sheet(input_file_abs_path, ws_history, df_master_data)
        else:
            print(Fore.RED, 'No stock data present for the given date')

    except Exception as excp:
        if PRINT_DETAILED_EXCEPTION:
            tb.print_exc()
        else:
            print(Fore.RED, f'\nException occured: {excp.__str__()}')
